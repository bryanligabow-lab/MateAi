"""
Suscripciones_Rewards_v2.xlsx — Sistema completo multi-hoja con control de pagos
"""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "/Users/Bryan/mate_ai/mateAIweb/Suscripciones_Rewards_v2.xlsx"
FONT = "Arial"

# ============== ESTILOS ==============
def font(**kw):
    kw.setdefault("name", FONT)
    return Font(**kw)

TITLE_FILL = PatternFill("solid", start_color="0F172A")
TITLE_FONT = font(bold=True, size=18, color="FFFFFF")
SUB_FONT   = font(italic=True, size=10, color="64748B")

H_FILL  = PatternFill("solid", start_color="1E3A8A")
H_FONT  = font(bold=True, size=10, color="FFFFFF")

INPUT_FONT   = font(size=10, color="1D4ED8")
FORMULA_FONT = font(size=10, color="111827")
NOTE_FONT    = font(italic=True, size=9, color="6B7280")
BOLD         = font(bold=True, size=10, color="111827")

THIN = Side(border_style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

# Status fills
GREEN  = PatternFill("solid", start_color="DCFCE7")
YELLOW = PatternFill("solid", start_color="FEF3C7")
RED    = PatternFill("solid", start_color="FECACA")
GRAY   = PatternFill("solid", start_color="E5E7EB")
BLUE_LIGHT = PatternFill("solid", start_color="DBEAFE")
KPI_VAL  = PatternFill("solid", start_color="FEF3C7")
KPI_LBL  = PatternFill("solid", start_color="F1F5F9")

wb = Workbook()
wb.remove(wb.active)

def new_sheet(name, widths=None, hide_gridlines=True):
    ws = wb.create_sheet(name)
    if hide_gridlines:
        ws.sheet_view.showGridLines = False
    if widths:
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
    return ws

def write_title(ws, title, subtitle, span=12):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=title)
    c.fill = TITLE_FILL
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 40
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = SUB_FONT
    c.alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[2].height = 18

def write_headers(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = H_FILL
        c.font = H_FONT
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 30

def style_row(ws, row, ncols, formula_cols=()):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.border = BORDER
        if col in formula_cols and c.value:
            c.font = FORMULA_FONT
        elif c.font.color is None:
            c.font = INPUT_FONT
        if c.alignment.horizontal is None:
            c.alignment = LEFT

# =================================================================
# HOJA 1 — PLANES
# =================================================================
ws_planes = new_sheet("Planes", widths={"A":14,"B":18,"C":14,"D":12,"E":18,"F":48,"G":10})
write_title(ws_planes, "📋 PLANES", "Catálogo maestro de planes de suscripción", span=7)
write_headers(ws_planes, 4, ["ID Plan","Nombre","Periodicidad","Precio","Puntos por ciclo","Beneficios incluidos","Activo"])

planes_data = [
    ("PLAN-BASIC",  "Basic",     "Mensual", 40,  50,   "Soporte estándar · 1 reward/mes",                   "Sí"),
    ("PLAN-PRO",    "Pro",       "Mensual", 80,  100,  "Soporte prioritario · 2 rewards · 5% desc. upsell", "Sí"),
    ("PLAN-ELITE",  "Elite",     "Mensual", 150, 200,  "Soporte 24/7 · rewards ilimitados · 10% desc.",     "Sí"),
    ("PLAN-ANUAL",  "Pro Anual", "Anual",   800, 1500, "Plan Pro pagado anual · 2 meses gratis",            "Sí"),
]
for i, p in enumerate(planes_data):
    r = 5 + i
    for j, v in enumerate(p):
        c = ws_planes.cell(row=r, column=j+1, value=v)
        c.font = INPUT_FONT
        c.border = BORDER
        c.alignment = LEFT if j not in (3,4) else RIGHT
    ws_planes.cell(row=r, column=4).number_format = '"$"#,##0'
    ws_planes.cell(row=r, column=5).number_format = '#,##0'

# Filas reservadas
for r in range(9, 15):
    for col in range(1, 8):
        ws_planes.cell(row=r, column=col).border = BORDER

wb.defined_names["TBL_PLANES"] = DefinedName("TBL_PLANES", attr_text="Planes!$A$5:$G$14")

# =================================================================
# HOJA 2 — CLIENTES
# =================================================================
ws_cli = new_sheet("Clientes", widths={"A":12,"B":24,"C":26,"D":16,"E":14,"F":10,"G":36})
write_title(ws_cli, "👤 CLIENTES", "Maestro de clientes con suscripción activa o histórica", span=7)
write_headers(ws_cli, 4, ["ID Cliente","Nombre","Email","Teléfono","Fecha alta","Activo","Notas"])

clientes_data = [
    ("CLI-001", "Jhon Amaya",    "jhon@paranoia.com",   "+57 300 111 1111", datetime.date(2026, 2, 25), "Sí", "Cliente desde marzo 2026"),
    ("CLI-002", "Restfull",      "ops@restfull.com",    "+57 300 222 2222", datetime.date(2026, 4, 1),  "Sí", ""),
    ("CLI-003", "Juli Riofrío",  "juli@avenaza.com",    "+57 300 333 3333", datetime.date(2026, 1, 10), "Sí", "Premium"),
    ("CLI-004", "Jean Aguilar",  "jean@cortepiedra.co", "+57 300 444 4444", datetime.date(2026, 4, 18), "Sí", "Pago manual"),
    ("CLI-005", "Avenaza SA",    "facturacion@avenaza.com","+57 300 555 5555", datetime.date(2026, 1, 30), "Sí", "Plan anual"),
    ("CLI-006", "Demo Cliente",  "demo@mateai.com",     "+57 300 666 6666", datetime.date(2026, 4, 25), "Sí", "Trial extendido"),
]
for i, cli in enumerate(clientes_data):
    r = 5 + i
    for j, v in enumerate(cli):
        c = ws_cli.cell(row=r, column=j+1, value=v)
        c.font = INPUT_FONT
        c.border = BORDER
        c.alignment = LEFT
    ws_cli.cell(row=r, column=5).number_format = "yyyy-mm-dd"

for r in range(11, 30):
    for col in range(1, 8):
        ws_cli.cell(row=r, column=col).border = BORDER

wb.defined_names["TBL_CLIENTES"] = DefinedName("TBL_CLIENTES", attr_text="Clientes!$A$5:$G$30")

# =================================================================
# HOJA 3 — SUSCRIPCIONES
# Cols: A ID Sub | B ID Cliente | C Cliente(auto) | D ID Plan | E Plan(auto) | F Periodicidad(auto)
#       G Precio(auto) | H Fecha inicio | I Día cobro | J Estado | K Auto-renov | L Método pago | M Notas
# =================================================================
ws_sub = new_sheet("Suscripciones", widths={"A":12,"B":12,"C":22,"D":14,"E":18,"F":14,"G":12,"H":12,"I":10,"J":14,"K":12,"L":16,"M":30})
write_title(ws_sub, "📌 SUSCRIPCIONES", "Una fila por suscripción · campos auto-calculados desde Planes y Clientes", span=13)
write_headers(ws_sub, 4, ["ID Sub","ID Cliente","Cliente (auto)","ID Plan","Plan (auto)","Periodicidad (auto)","Precio (auto)","Fecha inicio","Día de cobro","Estado","Auto-renov","Método pago","Notas"])

suscripciones_data = [
    # ID Sub, ID Cli, ID Plan, Fecha inicio, Día cobro, Estado, Auto-renov, Método, Notas
    ("SUB-0001","CLI-001","PLAN-PRO",   datetime.date(2026,3,1),  1, "Activa","Sí","Transferencia",""),
    ("SUB-0002","CLI-002","PLAN-BASIC", datetime.date(2026,4,5),  5, "Activa","Sí","Transferencia",""),
    ("SUB-0003","CLI-003","PLAN-ELITE", datetime.date(2026,1,15),15, "Activa","Sí","Tarjeta","Premium"),
    ("SUB-0004","CLI-004","PLAN-PRO",   datetime.date(2026,4,20),20, "Activa","No","Efectivo","Cobro manual"),
    ("SUB-0005","CLI-005","PLAN-ANUAL", datetime.date(2026,2,1),  1, "Activa","Sí","Transferencia",""),
    ("SUB-0006","CLI-006","PLAN-BASIC", datetime.date(2026,4,25),25, "Pausada","No","Transferencia","Trial"),
]
for i, s in enumerate(suscripciones_data):
    r = 5 + i
    # Inputs
    ws_sub.cell(row=r, column=1, value=s[0]).font = INPUT_FONT
    ws_sub.cell(row=r, column=2, value=s[1]).font = INPUT_FONT
    ws_sub.cell(row=r, column=4, value=s[2]).font = INPUT_FONT
    cd = ws_sub.cell(row=r, column=8, value=s[3]); cd.font = INPUT_FONT; cd.number_format = "yyyy-mm-dd"
    ws_sub.cell(row=r, column=9, value=s[4]).font = INPUT_FONT
    ws_sub.cell(row=r, column=10, value=s[5]).font = INPUT_FONT
    ws_sub.cell(row=r, column=11, value=s[6]).font = INPUT_FONT
    ws_sub.cell(row=r, column=12, value=s[7]).font = INPUT_FONT
    ws_sub.cell(row=r, column=13, value=s[8]).font = INPUT_FONT
    # Fórmulas auto
    ws_sub.cell(row=r, column=3, value=f'=IFERROR(VLOOKUP(B{r},TBL_CLIENTES,2,FALSE),"")').font = FORMULA_FONT
    ws_sub.cell(row=r, column=5, value=f'=IFERROR(VLOOKUP(D{r},TBL_PLANES,2,FALSE),"")').font = FORMULA_FONT
    ws_sub.cell(row=r, column=6, value=f'=IFERROR(VLOOKUP(D{r},TBL_PLANES,3,FALSE),"")').font = FORMULA_FONT
    cg = ws_sub.cell(row=r, column=7, value=f'=IFERROR(VLOOKUP(D{r},TBL_PLANES,4,FALSE),"")')
    cg.font = FORMULA_FONT; cg.number_format = '"$"#,##0'

    for col in range(1, 14):
        c = ws_sub.cell(row=r, column=col)
        c.border = BORDER
        if c.alignment.horizontal is None:
            c.alignment = RIGHT if col == 7 else LEFT

# Filas reservadas con fórmulas precargadas
for r in range(11, 30):
    ws_sub.cell(row=r, column=3, value=f'=IFERROR(VLOOKUP(B{r},TBL_CLIENTES,2,FALSE),"")').font = FORMULA_FONT
    ws_sub.cell(row=r, column=5, value=f'=IFERROR(VLOOKUP(D{r},TBL_PLANES,2,FALSE),"")').font = FORMULA_FONT
    ws_sub.cell(row=r, column=6, value=f'=IFERROR(VLOOKUP(D{r},TBL_PLANES,3,FALSE),"")').font = FORMULA_FONT
    cg = ws_sub.cell(row=r, column=7, value=f'=IFERROR(VLOOKUP(D{r},TBL_PLANES,4,FALSE),"")')
    cg.font = FORMULA_FONT; cg.number_format = '"$"#,##0'
    for col in range(1, 14):
        ws_sub.cell(row=r, column=col).border = BORDER

# Validaciones
dv_plan = DataValidation(type="list", formula1="=Planes!$A$5:$A$14", allow_blank=True); dv_plan.add("D5:D30"); ws_sub.add_data_validation(dv_plan)
dv_cli  = DataValidation(type="list", formula1="=Clientes!$A$5:$A$30", allow_blank=True); dv_cli.add("B5:B30"); ws_sub.add_data_validation(dv_cli)
dv_estado_sub = DataValidation(type="list", formula1='"Activa,Pausada,Cancelada"', allow_blank=True); dv_estado_sub.add("J5:J30"); ws_sub.add_data_validation(dv_estado_sub)
dv_yn = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True); dv_yn.add("K5:K30"); ws_sub.add_data_validation(dv_yn)
dv_metodo = DataValidation(type="list", formula1='"Transferencia,Tarjeta,Efectivo,PayPal,Otro"', allow_blank=True); dv_metodo.add("L5:L30"); ws_sub.add_data_validation(dv_metodo)

# Formato condicional Estado
ws_sub.conditional_formatting.add("J5:J30", FormulaRule(formula=['$J5="Activa"'],     fill=GREEN,  font=font(bold=True, color="166534")))
ws_sub.conditional_formatting.add("J5:J30", FormulaRule(formula=['$J5="Pausada"'],    fill=YELLOW, font=font(bold=True, color="92400E")))
ws_sub.conditional_formatting.add("J5:J30", FormulaRule(formula=['$J5="Cancelada"'],  fill=GRAY,   font=font(bold=True, color="374151")))

wb.defined_names["TBL_SUB"] = DefinedName("TBL_SUB", attr_text="Suscripciones!$A$5:$M$30")

# =================================================================
# HOJA 4 — PAGOS
# Cols: A ID Pago | B Fecha pago | C ID Sub | D Cliente(auto) | E Plan(auto) | F Monto | G Método | H Período (texto) | I Notas
# =================================================================
ws_pag = new_sheet("Pagos", widths={"A":12,"B":12,"C":12,"D":22,"E":14,"F":12,"G":16,"H":18,"I":30})
write_title(ws_pag, "💳 PAGOS", "Histórico de cobros · cada fila es un pago realizado", span=9)
write_headers(ws_pag, 4, ["ID Pago","Fecha pago","ID Sub","Cliente (auto)","Plan (auto)","Monto","Método","Período cubierto","Notas"])

pagos_data = [
    ("PAG-0001", datetime.date(2026,3,1),  "SUB-0001", 80,  "Transferencia", "2026-03", ""),
    ("PAG-0002", datetime.date(2026,4,1),  "SUB-0001", 80,  "Transferencia", "2026-04", ""),
    ("PAG-0003", datetime.date(2026,4,5),  "SUB-0002", 40,  "Transferencia", "2026-04", ""),
    ("PAG-0004", datetime.date(2026,1,15), "SUB-0003", 150, "Tarjeta",       "2026-01", ""),
    ("PAG-0005", datetime.date(2026,2,15), "SUB-0003", 150, "Tarjeta",       "2026-02", ""),
    ("PAG-0006", datetime.date(2026,3,15), "SUB-0003", 150, "Tarjeta",       "2026-03", ""),
    ("PAG-0007", datetime.date(2026,4,15), "SUB-0003", 150, "Tarjeta",       "2026-04", ""),
    ("PAG-0008", datetime.date(2026,4,20), "SUB-0004", 80,  "Efectivo",      "2026-04", ""),
    ("PAG-0009", datetime.date(2026,2,1),  "SUB-0005", 800, "Transferencia", "2026 anual", "Plan anual completo"),
]
for i, p in enumerate(pagos_data):
    r = 5 + i
    ws_pag.cell(row=r, column=1, value=p[0]).font = INPUT_FONT
    cb = ws_pag.cell(row=r, column=2, value=p[1]); cb.font = INPUT_FONT; cb.number_format = "yyyy-mm-dd"
    ws_pag.cell(row=r, column=3, value=p[2]).font = INPUT_FONT
    cf = ws_pag.cell(row=r, column=6, value=p[3]); cf.font = INPUT_FONT; cf.number_format = '"$"#,##0'
    ws_pag.cell(row=r, column=7, value=p[4]).font = INPUT_FONT
    ws_pag.cell(row=r, column=8, value=p[5]).font = INPUT_FONT
    ws_pag.cell(row=r, column=9, value=p[6]).font = INPUT_FONT
    # Auto: Cliente y Plan
    ws_pag.cell(row=r, column=4, value=f'=IFERROR(VLOOKUP(C{r},TBL_SUB,3,FALSE),"")').font = FORMULA_FONT
    ws_pag.cell(row=r, column=5, value=f'=IFERROR(VLOOKUP(C{r},TBL_SUB,5,FALSE),"")').font = FORMULA_FONT
    for col in range(1, 10):
        c = ws_pag.cell(row=r, column=col); c.border = BORDER
        if c.alignment.horizontal is None: c.alignment = RIGHT if col==6 else LEFT

# Filas reservadas
for r in range(14, 50):
    ws_pag.cell(row=r, column=4, value=f'=IFERROR(VLOOKUP(C{r},TBL_SUB,3,FALSE),"")').font = FORMULA_FONT
    ws_pag.cell(row=r, column=5, value=f'=IFERROR(VLOOKUP(C{r},TBL_SUB,5,FALSE),"")').font = FORMULA_FONT
    for col in range(1, 10):
        ws_pag.cell(row=r, column=col).border = BORDER
    ws_pag.cell(row=r, column=2).number_format = "yyyy-mm-dd"
    ws_pag.cell(row=r, column=6).number_format = '"$"#,##0'

dv_sub = DataValidation(type="list", formula1="=Suscripciones!$A$5:$A$30", allow_blank=True); dv_sub.add("C5:C50"); ws_pag.add_data_validation(dv_sub)
dv_metodo2 = DataValidation(type="list", formula1='"Transferencia,Tarjeta,Efectivo,PayPal,Otro"', allow_blank=True); dv_metodo2.add("G5:G50"); ws_pag.add_data_validation(dv_metodo2)

wb.defined_names["TBL_PAGOS"] = DefinedName("TBL_PAGOS", attr_text="Pagos!$A$5:$I$50")

# =================================================================
# HOJA 5 — ESTADO DE CUENTA (auto-calculada desde Suscripciones + Pagos)
# Cols: A ID Sub | B Cliente | C Plan | D Periodicidad | E Precio | F Inicio | G Día cobro
#       H Cuotas vencidas (esperadas) | I Total facturado | J Cuotas pagadas | K Total pagado
#       L Saldo pendiente | M Próximo cobro | N Días para cobro | O Estado pago
# =================================================================
ws_ec = new_sheet("EstadoCuenta", widths={"A":12,"B":22,"C":16,"D":14,"E":12,"F":12,"G":10,"H":14,"I":14,"J":14,"K":14,"L":14,"M":14,"N":14,"O":18})
write_title(ws_ec, "🧾 ESTADO DE CUENTA", "Saldo, mora y próximo cobro por suscripción · todo automático desde Suscripciones + Pagos", span=15)
write_headers(ws_ec, 4, ["ID Sub","Cliente","Plan","Periodicidad","Precio","Inicio","Día cobro",
                          "Cuotas esperadas","Total facturado","Cuotas pagadas","Total pagado",
                          "Saldo pendiente","Próximo cobro","Días para cobro","Estado pago"])

# Fórmulas: una fila por cada SUB-#### posible (referencia a Suscripciones por fila)
# Vamos a referenciar Suscripciones directamente fila a fila para mantenerlo simple y dinámico.
for idx in range(26):  # 26 filas posibles (alineadas con SUB rows 5:30)
    sub_r = 5 + idx     # fila en Suscripciones
    r = 5 + idx         # fila en EstadoCuenta

    # A: ID Sub (mirror)
    ws_ec.cell(row=r, column=1, value=f"=IF(Suscripciones!A{sub_r}=\"\",\"\",Suscripciones!A{sub_r})").font = FORMULA_FONT
    # B: Cliente
    ws_ec.cell(row=r, column=2, value=f"=IF(Suscripciones!A{sub_r}=\"\",\"\",Suscripciones!C{sub_r})").font = FORMULA_FONT
    # C: Plan
    ws_ec.cell(row=r, column=3, value=f"=IF(Suscripciones!A{sub_r}=\"\",\"\",Suscripciones!E{sub_r})").font = FORMULA_FONT
    # D: Periodicidad
    ws_ec.cell(row=r, column=4, value=f"=IF(Suscripciones!A{sub_r}=\"\",\"\",Suscripciones!F{sub_r})").font = FORMULA_FONT
    # E: Precio
    ce = ws_ec.cell(row=r, column=5, value=f"=IF(Suscripciones!A{sub_r}=\"\",\"\",Suscripciones!G{sub_r})")
    ce.font = FORMULA_FONT; ce.number_format = '"$"#,##0;[Red]-"$"#,##0;-'
    # F: Inicio
    cf = ws_ec.cell(row=r, column=6, value=f"=IF(Suscripciones!A{sub_r}=\"\",\"\",Suscripciones!H{sub_r})")
    cf.font = FORMULA_FONT; cf.number_format = "yyyy-mm-dd"
    # G: Día cobro
    ws_ec.cell(row=r, column=7, value=f"=IF(Suscripciones!A{sub_r}=\"\",\"\",Suscripciones!I{sub_r})").font = FORMULA_FONT

    # H: Cuotas esperadas hasta hoy
    # Si Estado = Cancelada o Pausada -> 0 nuevas cuotas (mantenemos lo que ya debía hasta esa fecha — simplificado: 0)
    # Mensual: meses transcurridos desde inicio +1 (incluye el mes de inicio)
    # Anual: años transcurridos +1
    h_formula = (
        f'=IF(Suscripciones!A{sub_r}="","",'
        f'IF(Suscripciones!J{sub_r}<>"Activa",0,'
        f'IF(Suscripciones!F{sub_r}="Anual",'
        f'  MAX(0,DATEDIF(Suscripciones!H{sub_r},TODAY(),"Y")+1),'
        f'  MAX(0,DATEDIF(Suscripciones!H{sub_r},TODAY(),"M")+1)'
        f')))'
    )
    ws_ec.cell(row=r, column=8, value=h_formula).font = FORMULA_FONT

    # I: Total facturado = cuotas esperadas * precio
    ci = ws_ec.cell(row=r, column=9, value=f'=IF(A{r}="","",H{r}*E{r})')
    ci.font = FORMULA_FONT; ci.number_format = '"$"#,##0;[Red]-"$"#,##0;-'

    # J: Cuotas pagadas
    cj = ws_ec.cell(row=r, column=10, value=f'=IF(A{r}="","",COUNTIFS(Pagos!C:C,A{r}))')
    cj.font = FORMULA_FONT; cj.number_format = '#,##0;-#,##0;-'

    # K: Total pagado
    ck = ws_ec.cell(row=r, column=11, value=f'=IF(A{r}="","",SUMIFS(Pagos!F:F,Pagos!C:C,A{r}))')
    ck.font = FORMULA_FONT; ck.number_format = '"$"#,##0;[Red]-"$"#,##0;-'

    # L: Saldo pendiente = facturado - pagado
    cl = ws_ec.cell(row=r, column=12, value=f'=IF(A{r}="","",I{r}-K{r})')
    cl.font = FORMULA_FONT; cl.number_format = '"$"#,##0;[Red]-"$"#,##0;-'

    # M: Próximo cobro
    # Si periodicidad mensual: EDATE(inicio, cuotas_pagadas) -> siguiente fecha = EDATE(inicio, cuotas_pagadas)
    # Si periodicidad anual: DATE(year(inicio)+cuotas_pagadas, month(inicio), day(inicio))
    m_formula = (
        f'=IF(A{r}="","",'
        f'IF(Suscripciones!J{sub_r}<>"Activa","",'
        f'IF(D{r}="Anual",'
        f'  DATE(YEAR(F{r})+J{r},MONTH(F{r}),DAY(F{r})),'
        f'  EDATE(F{r},J{r})'
        f')))'
    )
    cm = ws_ec.cell(row=r, column=13, value=m_formula)
    cm.font = FORMULA_FONT; cm.number_format = "yyyy-mm-dd"

    # N: Días para cobro = M - TODAY()
    cn = ws_ec.cell(row=r, column=14, value=f'=IF(OR(A{r}="",M{r}=""),"",M{r}-TODAY())')
    cn.font = FORMULA_FONT; cn.number_format = '0;[Red]-0;-'

    # O: Estado pago
    # Si sub no activa -> "Pausada/Cancelada"
    # Si saldo > 0 y N < 0 -> "En mora"
    # Si saldo > 0 y N entre 0 y 7 -> "Por cobrar"
    # Si saldo > 0 y N > 7 -> "Próximo"
    # Si saldo <= 0 -> "Al día"
    o_formula = (
        f'=IF(A{r}="","",'
        f'IF(Suscripciones!J{sub_r}<>"Activa",Suscripciones!J{sub_r},'
        f'IF(L{r}<=0,"Al día",'
        f'IF(N{r}<0,"En mora",'
        f'IF(N{r}<=7,"Por cobrar","Próximo")))))'
    )
    ws_ec.cell(row=r, column=15, value=o_formula).font = FORMULA_FONT

    for col in range(1, 16):
        c = ws_ec.cell(row=r, column=col); c.border = BORDER
        if c.alignment.horizontal is None:
            c.alignment = RIGHT if col in (5,7,8,9,10,11,12,14) else LEFT

# Formato condicional Estado pago (col O)
ws_ec.conditional_formatting.add("O5:O30", FormulaRule(formula=['$O5="Al día"'],       fill=GREEN,  font=font(bold=True, color="166534")))
ws_ec.conditional_formatting.add("O5:O30", FormulaRule(formula=['$O5="Por cobrar"'],   fill=YELLOW, font=font(bold=True, color="92400E")))
ws_ec.conditional_formatting.add("O5:O30", FormulaRule(formula=['$O5="En mora"'],      fill=RED,    font=font(bold=True, color="991B1B")))
ws_ec.conditional_formatting.add("O5:O30", FormulaRule(formula=['$O5="Próximo"'],      fill=BLUE_LIGHT, font=font(bold=True, color="1E3A8A")))
ws_ec.conditional_formatting.add("O5:O30", FormulaRule(formula=['$O5="Pausada"'],      fill=GRAY,   font=font(bold=True, color="374151")))
ws_ec.conditional_formatting.add("O5:O30", FormulaRule(formula=['$O5="Cancelada"'],    fill=GRAY,   font=font(bold=True, color="374151")))

# Saldo pendiente: rojo si > 0
ws_ec.conditional_formatting.add("L5:L30", FormulaRule(formula=['AND($L5<>"",$L5>0)'], fill=RED, font=font(bold=True, color="991B1B")))
ws_ec.conditional_formatting.add("L5:L30", FormulaRule(formula=['AND($L5<>"",$L5<=0)'], fill=GREEN, font=font(bold=True, color="166534")))

wb.defined_names["TBL_EC"] = DefinedName("TBL_EC", attr_text="EstadoCuenta!$A$5:$O$30")

# =================================================================
# HOJA 6 — REWARDS (catálogo)
# =================================================================
ws_rw = new_sheet("Rewards", widths={"A":12,"B":24,"C":14,"D":14,"E":14,"F":36,"G":10})
write_title(ws_rw, "🎁 REWARDS", "Catálogo de beneficios canjeables con puntos", span=7)
write_headers(ws_rw, 4, ["ID Reward","Nombre","Tipo","Costo puntos","Plan mínimo","Descripción","Activo"])

rewards_data = [
    ("RW-001","Descuento 10%",        "Descuento",   100, "PLAN-BASIC","10% off en próximo upsell",       "Sí"),
    ("RW-002","Descuento 25%",        "Descuento",   250, "PLAN-PRO",  "25% off en próximo upsell",       "Sí"),
    ("RW-003","Mes gratis",           "Suscripción", 500, "PLAN-PRO",  "1 mes gratis al renovar",         "Sí"),
    ("RW-004","Consultoría 30 min",   "Servicio",    300, "PLAN-PRO",  "Sesión 1:1 con especialista",     "Sí"),
    ("RW-005","Consultoría 60 min",   "Servicio",    600, "PLAN-ELITE","Sesión estratégica 60 min",       "Sí"),
    ("RW-006","Acceso early features","Acceso",      150, "PLAN-BASIC","Beta de nuevas funciones",        "Sí"),
    ("RW-007","Branding pack",        "Producto",    400, "PLAN-PRO",  "Plantillas de marca exclusivas",  "Sí"),
    ("RW-008","Auditoría completa",   "Servicio",   1000, "PLAN-ELITE","Auditoría de negocio integral",   "Sí"),
]
for i, rw in enumerate(rewards_data):
    r = 5 + i
    for j, v in enumerate(rw):
        c = ws_rw.cell(row=r, column=j+1, value=v)
        c.font = INPUT_FONT
        c.border = BORDER
        c.alignment = LEFT if j != 3 else RIGHT
    ws_rw.cell(row=r, column=4).number_format = "#,##0"

for r in range(13, 22):
    for col in range(1, 8):
        ws_rw.cell(row=r, column=col).border = BORDER

wb.defined_names["TBL_REWARDS"] = DefinedName("TBL_REWARDS", attr_text="Rewards!$A$5:$G$22")

# =================================================================
# HOJA 7 — CANJES
# Cols: A ID Canje | B Fecha | C ID Cliente | D Cliente(auto) | E ID Sub | F ID Reward | G Reward(auto) | H Costo puntos(auto) | I Origen | J Estado
# =================================================================
ws_cj = new_sheet("Canjes", widths={"A":12,"B":12,"C":12,"D":22,"E":12,"F":12,"G":24,"H":14,"I":16,"J":14})
write_title(ws_cj, "🏆 CANJES", "Registro de rewards solicitados/canjeados por cliente", span=10)
write_headers(ws_cj, 4, ["ID Canje","Fecha","ID Cliente","Cliente (auto)","ID Sub","ID Reward","Reward (auto)","Costo puntos (auto)","Origen","Estado"])

canjes_data = [
    ("CJ-0001", datetime.date(2026,4,10), "CLI-001", "SUB-0001", "RW-001", "Suscripción", "Canjeado"),
    ("CJ-0002", datetime.date(2026,4,15), "CLI-003", "SUB-0003", "RW-005", "Suscripción", "Canjeado"),
    ("CJ-0003", datetime.date(2026,4,22), "CLI-003", "SUB-0003", "RW-007", "Manual",       "Canjeado"),
    ("CJ-0004", datetime.date(2026,4,28), "CLI-001", "SUB-0001", "RW-006", "Automático",  "Pendiente"),
    ("CJ-0005", datetime.date(2026,5,1),  "CLI-005", "SUB-0005", "RW-002", "Suscripción", "Canjeado"),
]
for i, x in enumerate(canjes_data):
    r = 5 + i
    ws_cj.cell(row=r, column=1, value=x[0]).font = INPUT_FONT
    cb = ws_cj.cell(row=r, column=2, value=x[1]); cb.font = INPUT_FONT; cb.number_format = "yyyy-mm-dd"
    ws_cj.cell(row=r, column=3, value=x[2]).font = INPUT_FONT
    ws_cj.cell(row=r, column=5, value=x[3]).font = INPUT_FONT
    ws_cj.cell(row=r, column=6, value=x[4]).font = INPUT_FONT
    ws_cj.cell(row=r, column=9, value=x[5]).font = INPUT_FONT
    ws_cj.cell(row=r, column=10, value=x[6]).font = INPUT_FONT
    # Auto
    ws_cj.cell(row=r, column=4, value=f'=IFERROR(VLOOKUP(C{r},TBL_CLIENTES,2,FALSE),"")').font = FORMULA_FONT
    ws_cj.cell(row=r, column=7, value=f'=IFERROR(VLOOKUP(F{r},TBL_REWARDS,2,FALSE),"")').font = FORMULA_FONT
    ch = ws_cj.cell(row=r, column=8, value=f'=IFERROR(VLOOKUP(F{r},TBL_REWARDS,4,FALSE),"")')
    ch.font = FORMULA_FONT; ch.number_format = "#,##0"
    for col in range(1, 11):
        c = ws_cj.cell(row=r, column=col); c.border = BORDER
        if c.alignment.horizontal is None: c.alignment = RIGHT if col == 8 else LEFT

for r in range(10, 30):
    ws_cj.cell(row=r, column=4, value=f'=IFERROR(VLOOKUP(C{r},TBL_CLIENTES,2,FALSE),"")').font = FORMULA_FONT
    ws_cj.cell(row=r, column=7, value=f'=IFERROR(VLOOKUP(F{r},TBL_REWARDS,2,FALSE),"")').font = FORMULA_FONT
    ch = ws_cj.cell(row=r, column=8, value=f'=IFERROR(VLOOKUP(F{r},TBL_REWARDS,4,FALSE),"")')
    ch.font = FORMULA_FONT; ch.number_format = "#,##0"
    for col in range(1, 11):
        ws_cj.cell(row=r, column=col).border = BORDER
    ws_cj.cell(row=r, column=2).number_format = "yyyy-mm-dd"

dv_estcj = DataValidation(type="list", formula1='"Pendiente,Canjeado,Expirado,Cancelado"', allow_blank=True); dv_estcj.add("J5:J30"); ws_cj.add_data_validation(dv_estcj)
dv_orig  = DataValidation(type="list", formula1='"Suscripción,Manual,Automático,Promoción"', allow_blank=True); dv_orig.add("I5:I30"); ws_cj.add_data_validation(dv_orig)
dv_rw    = DataValidation(type="list", formula1="=Rewards!$A$5:$A$22", allow_blank=True); dv_rw.add("F5:F30"); ws_cj.add_data_validation(dv_rw)
dv_cli2  = DataValidation(type="list", formula1="=Clientes!$A$5:$A$30", allow_blank=True); dv_cli2.add("C5:C30"); ws_cj.add_data_validation(dv_cli2)

ws_cj.conditional_formatting.add("J5:J30", FormulaRule(formula=['$J5="Canjeado"'],  fill=GREEN,  font=font(bold=True, color="166534")))
ws_cj.conditional_formatting.add("J5:J30", FormulaRule(formula=['$J5="Pendiente"'], fill=YELLOW, font=font(bold=True, color="92400E")))
ws_cj.conditional_formatting.add("J5:J30", FormulaRule(formula=['$J5="Expirado"'],  fill=RED,    font=font(bold=True, color="991B1B")))
ws_cj.conditional_formatting.add("J5:J30", FormulaRule(formula=['$J5="Cancelado"'], fill=GRAY,   font=font(bold=True, color="374151")))

# =================================================================
# HOJA 8 — ALERTAS (vista filtrada)
# =================================================================
ws_al = new_sheet("Alertas", widths={"A":18,"B":12,"C":22,"D":14,"E":14,"F":12,"G":14,"H":14,"I":18})
write_title(ws_al, "🚨 ALERTAS DE COBRO", "Suscripciones por cobrar próximamente y clientes en mora · auto desde EstadoCuenta", span=9)

# Sección 1: Por cobrar (próximos 7 días)
ws_al.cell(row=4, column=1, value="▼ POR COBRAR (próximos 7 días)").font = font(bold=True, size=12, color="92400E")
ws_al.merge_cells("A4:I4")
write_headers(ws_al, 5, ["Estado","ID Sub","Cliente","Plan","Precio","Próx. cobro","Días para cobro","Saldo","Método pago"])

# Usamos FILTER si está disponible (Excel 365 / Sheets) — y como fallback referencias directas
# Vamos a poner una fórmula maestra con FILTER + alternativa con IFERROR
filter_por_cobrar = (
    '=IFERROR(SORT(FILTER(EstadoCuenta!O5:O30&"|"&EstadoCuenta!A5:O30,EstadoCuenta!O5:O30="Por cobrar"),0,1),"")'
)
# Mejor: 12 filas con INDEX por número de orden
for k in range(12):
    r = 6 + k
    # Cargamos fila k-ésima donde Estado="Por cobrar"
    # Usamos IFERROR(INDEX(SMALL(...))) approach
    # Para mantenerlo simple y compatible: usamos INDEX+MATCH+SMALL
    # k+1-ésima coincidencia
    occurrence = k + 1
    base_match = (
        f'IFERROR(SMALL(IF(EstadoCuenta!$O$5:$O$30="Por cobrar",ROW(EstadoCuenta!$O$5:$O$30)),{occurrence}),"")'
    )
    # En todas las celdas usamos esa misma fila
    cols_map = [
        ("A", 'EstadoCuenta!$O', None),    # Estado
        ("B", 'EstadoCuenta!$A', None),    # ID Sub
        ("C", 'EstadoCuenta!$B', None),    # Cliente
        ("D", 'EstadoCuenta!$C', None),    # Plan
        ("E", 'EstadoCuenta!$E', '"$"#,##0'),  # Precio
        ("F", 'EstadoCuenta!$M', "yyyy-mm-dd"),  # Próx cobro
        ("G", 'EstadoCuenta!$N', '0;[Red]-0;-'), # Días
        ("H", 'EstadoCuenta!$L', '"$"#,##0'),   # Saldo
    ]
    for col_letter, ref_col, num_fmt in cols_map:
        f = f'=IFERROR(INDEX({ref_col}:{ref_col[-2:]}$30,{base_match}-4),"")'
        # ref_col like 'EstadoCuenta!$A' -> we want $A:$A but with ROW reference compensated
        # Actually we need INDEX(EstadoCuenta!$A$5:$A$30, position)
        # base_match returns ROW # (5..30); position = ROW#-4
        f = f'=IFERROR(INDEX({ref_col}$5:{ref_col[-2:]}$30,{base_match}-4),"")'
        c = ws_al[f"{col_letter}{r}"]
        c.value = f
        c.font = FORMULA_FONT
        c.border = BORDER
        if num_fmt:
            c.number_format = num_fmt
    # Método pago: lookup desde Suscripciones
    fm = f'=IFERROR(VLOOKUP(B{r},Suscripciones!$A$5:$L$30,12,FALSE),"")'
    cm = ws_al[f"I{r}"]
    cm.value = fm; cm.font = FORMULA_FONT; cm.border = BORDER
    cm.alignment = LEFT
    # Bordes y alineación
    for col_letter in ["A","B","C","D","E","F","G","H","I"]:
        cc = ws_al[f"{col_letter}{r}"]
        cc.border = BORDER
        if col_letter in ("E","G","H"):
            cc.alignment = RIGHT
        else:
            cc.alignment = LEFT

# Sección 2: En mora
mora_start = 19
ws_al.cell(row=mora_start, column=1, value="▼ EN MORA").font = font(bold=True, size=12, color="991B1B")
ws_al.merge_cells(f"A{mora_start}:I{mora_start}")
write_headers(ws_al, mora_start+1, ["Estado","ID Sub","Cliente","Plan","Precio","Próx. cobro","Días mora","Saldo","Método pago"])

for k in range(12):
    r = mora_start + 2 + k
    occurrence = k + 1
    base_match = (
        f'IFERROR(SMALL(IF(EstadoCuenta!$O$5:$O$30="En mora",ROW(EstadoCuenta!$O$5:$O$30)),{occurrence}),"")'
    )
    cols_map = [
        ("A", 'EstadoCuenta!$O', None),
        ("B", 'EstadoCuenta!$A', None),
        ("C", 'EstadoCuenta!$B', None),
        ("D", 'EstadoCuenta!$C', None),
        ("E", 'EstadoCuenta!$E', '"$"#,##0'),
        ("F", 'EstadoCuenta!$M', "yyyy-mm-dd"),
        ("G", 'EstadoCuenta!$N', '0;[Red]-0;-'),
        ("H", 'EstadoCuenta!$L', '"$"#,##0'),
    ]
    for col_letter, ref_col, num_fmt in cols_map:
        f = f'=IFERROR(INDEX({ref_col}$5:{ref_col[-2:]}$30,{base_match}-4),"")'
        c = ws_al[f"{col_letter}{r}"]
        c.value = f
        c.font = FORMULA_FONT
        c.border = BORDER
        if num_fmt:
            c.number_format = num_fmt
    fm = f'=IFERROR(VLOOKUP(B{r},Suscripciones!$A$5:$L$30,12,FALSE),"")'
    cm = ws_al[f"I{r}"]
    cm.value = fm; cm.font = FORMULA_FONT; cm.border = BORDER
    for col_letter in ["A","B","C","D","E","F","G","H","I"]:
        cc = ws_al[f"{col_letter}{r}"]
        cc.border = BORDER
        if col_letter in ("E","G","H"):
            cc.alignment = RIGHT
        else:
            cc.alignment = LEFT

# Formato condicional para alertas
ws_al.conditional_formatting.add("A6:A17", FormulaRule(formula=['$A6="Por cobrar"'], fill=YELLOW, font=font(bold=True, color="92400E")))
ws_al.conditional_formatting.add(f"A{mora_start+2}:A{mora_start+13}", FormulaRule(formula=[f'$A{mora_start+2}="En mora"'], fill=RED, font=font(bold=True, color="991B1B")))

# =================================================================
# HOJA 9 — DASHBOARD (KPIs)
# =================================================================
ws_db = new_sheet("Dashboard", widths={"A":36,"B":18,"C":50})
# Título grande
ws_db.merge_cells("A1:C1")
t = ws_db.cell(row=1, column=1, value="📊 DASHBOARD — SUSCRIPCIONES & REWARDS")
t.fill = TITLE_FILL; t.font = TITLE_FONT
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_db.row_dimensions[1].height = 44
ws_db.merge_cells("A2:C2")
s = ws_db.cell(row=2, column=1, value=f"Actualizado al abrir · TODAY() recalcula automáticamente · MateAI")
s.font = SUB_FONT; s.alignment = Alignment(horizontal="left", indent=1)

# Sección INGRESOS
ws_db.merge_cells("A4:C4")
ws_db.cell(row=4, column=1, value="💰 INGRESOS").font = font(bold=True, size=13, color="0F172A")
ws_db.cell(row=4, column=1).fill = PatternFill("solid", start_color="DBEAFE")
ws_db.row_dimensions[4].height = 26

write_headers(ws_db, 5, ["KPI", "Valor", "Detalle"])

kpis_ingresos = [
    ("MRR (Ingreso recurrente mensual)",
     '=SUMPRODUCT((Suscripciones!F5:F30="Mensual")*(Suscripciones!J5:J30="Activa")*Suscripciones!G5:G30)+SUMPRODUCT((Suscripciones!F5:F30="Anual")*(Suscripciones!J5:J30="Activa")*Suscripciones!G5:G30)/12',
     "Mensuales activos + (Anuales activos / 12)", '"$"#,##0'),
    ("ARR (Ingreso anualizado)",
     '=B6*12',
     "MRR × 12", '"$"#,##0'),
    ("Cobrado este mes",
     '=SUMIFS(Pagos!F:F,Pagos!B:B,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Pagos!B:B,"<"&DATE(YEAR(TODAY()),MONTH(TODAY())+1,1))',
     "Suma de pagos del mes actual", '"$"#,##0'),
    ("Cobrado total histórico",
     '=SUM(Pagos!F:F)',
     "Suma de todos los pagos registrados", '"$"#,##0'),
    ("Saldo pendiente total",
     '=SUMIF(EstadoCuenta!L5:L30,">0")',
     "Suma de saldos positivos (clientes que deben)", '"$"#,##0'),
]
for i, (label, formula, detail, fmt) in enumerate(kpis_ingresos):
    r = 6 + i
    cl = ws_db.cell(row=r, column=1, value=label); cl.fill = KPI_LBL; cl.font = BOLD; cl.alignment = LEFT; cl.border = BORDER
    cv = ws_db.cell(row=r, column=2, value=formula); cv.fill = KPI_VAL; cv.font = font(bold=True, size=12, color="0F172A"); cv.alignment = RIGHT; cv.border = BORDER
    if fmt: cv.number_format = fmt
    cd = ws_db.cell(row=r, column=3, value=detail); cd.font = NOTE_FONT; cd.alignment = LEFT; cd.border = BORDER
    ws_db.row_dimensions[r].height = 24

# Sección SUSCRIPCIONES
ws_db.merge_cells("A12:C12")
ws_db.cell(row=12, column=1, value="📌 SUSCRIPCIONES").font = font(bold=True, size=13, color="0F172A")
ws_db.cell(row=12, column=1).fill = PatternFill("solid", start_color="DBEAFE")
ws_db.row_dimensions[12].height = 26
write_headers(ws_db, 13, ["KPI", "Valor", "Detalle"])

kpis_subs = [
    ("Suscripciones activas",
     '=COUNTIF(Suscripciones!J5:J30,"Activa")',
     "Estado = Activa", "#,##0"),
    ("Suscripciones pausadas",
     '=COUNTIF(Suscripciones!J5:J30,"Pausada")',
     "Estado = Pausada", "#,##0"),
    ("Suscripciones canceladas",
     '=COUNTIF(Suscripciones!J5:J30,"Cancelada")',
     "Estado = Cancelada", "#,##0"),
    ("Clientes al día",
     '=COUNTIF(EstadoCuenta!O5:O30,"Al día")',
     "Sin saldo pendiente", "#,##0"),
    ("Clientes por cobrar (≤7 días)",
     '=COUNTIF(EstadoCuenta!O5:O30,"Por cobrar")',
     "Próximo cobro en los próximos 7 días", "#,##0"),
    ("Clientes en mora",
     '=COUNTIF(EstadoCuenta!O5:O30,"En mora")',
     "Saldo pendiente y fecha de cobro vencida", "#,##0"),
    ("Plan más popular",
     '=IFERROR(INDEX(Suscripciones!E5:E30,MATCH(MAX(COUNTIF(Suscripciones!E5:E30,Suscripciones!E5:E30)),COUNTIF(Suscripciones!E5:E30,Suscripciones!E5:E30),0)),"")',
     "Plan con más suscripciones", None),
]
for i, (label, formula, detail, fmt) in enumerate(kpis_subs):
    r = 14 + i
    cl = ws_db.cell(row=r, column=1, value=label); cl.fill = KPI_LBL; cl.font = BOLD; cl.alignment = LEFT; cl.border = BORDER
    cv = ws_db.cell(row=r, column=2, value=formula); cv.fill = KPI_VAL; cv.font = font(bold=True, size=12, color="0F172A"); cv.alignment = RIGHT; cv.border = BORDER
    if fmt: cv.number_format = fmt
    cd = ws_db.cell(row=r, column=3, value=detail); cd.font = NOTE_FONT; cd.alignment = LEFT; cd.border = BORDER
    ws_db.row_dimensions[r].height = 24

# Formato especial para "Clientes en mora" — rojo si > 0
ws_db.conditional_formatting.add("B19:B19", FormulaRule(formula=['$B$19>0'], fill=RED, font=font(bold=True, size=12, color="991B1B")))
ws_db.conditional_formatting.add("B10:B10", FormulaRule(formula=['$B$10>0'], fill=RED, font=font(bold=True, size=12, color="991B1B")))

# Sección REWARDS
ws_db.merge_cells("A22:C22")
ws_db.cell(row=22, column=1, value="🎁 REWARDS").font = font(bold=True, size=13, color="0F172A")
ws_db.cell(row=22, column=1).fill = PatternFill("solid", start_color="DBEAFE")
ws_db.row_dimensions[22].height = 26
write_headers(ws_db, 23, ["KPI", "Valor", "Detalle"])

# Puntos disponibles por cliente = puntos generados por sub - puntos gastados en canjes (estado canjeado)
# Total puntos otorgados = SUMPRODUCT por sub: cuotas pagadas (col J EstadoCuenta) * puntos por ciclo del plan
# Más simple: para cada SUB calculamos cuotas pagadas * puntos del plan
kpis_rw = [
    ("Total puntos otorgados",
     '=SUMPRODUCT((EstadoCuenta!A5:A30<>"")*EstadoCuenta!J5:J30*IFERROR(VLOOKUP(EstadoCuenta!C5:C30,Planes!$B$5:$E$14,4,FALSE),0))',
     "Cuotas pagadas × puntos por ciclo del plan", "#,##0"),
    ("Puntos gastados (canjeados)",
     '=SUMIF(Canjes!J5:J30,"Canjeado",Canjes!H5:H30)',
     "Suma de costo en puntos de canjes confirmados", "#,##0"),
    ("Saldo de puntos disponibles",
     '=B24-B25',
     "Otorgados − gastados", "#,##0"),
    ("Total canjes (todos)",
     '=COUNTA(Canjes!A5:A30)',
     "Filas registradas en Canjes", "#,##0"),
    ("Canjes pendientes",
     '=COUNTIF(Canjes!J5:J30,"Pendiente")',
     "Estado = Pendiente (acción requerida)", "#,##0"),
    ("Reward más canjeado",
     '=IFERROR(INDEX(Canjes!G5:G30,MATCH(MAX(COUNTIF(Canjes!G5:G30,Canjes!G5:G30)),COUNTIF(Canjes!G5:G30,Canjes!G5:G30),0)),"")',
     "Reward con más canjes", None),
]
for i, (label, formula, detail, fmt) in enumerate(kpis_rw):
    r = 24 + i
    cl = ws_db.cell(row=r, column=1, value=label); cl.fill = KPI_LBL; cl.font = BOLD; cl.alignment = LEFT; cl.border = BORDER
    cv = ws_db.cell(row=r, column=2, value=formula); cv.fill = KPI_VAL; cv.font = font(bold=True, size=12, color="0F172A"); cv.alignment = RIGHT; cv.border = BORDER
    if fmt: cv.number_format = fmt
    cd = ws_db.cell(row=r, column=3, value=detail); cd.font = NOTE_FONT; cd.alignment = LEFT; cd.border = BORDER
    ws_db.row_dimensions[r].height = 24

# Notas finales
ws_db.row_dimensions[31].height = 8
ws_db.merge_cells("A32:C32")
n = ws_db.cell(row=32, column=1, value="📌 Cómo usar el sistema")
n.font = font(bold=True, size=12, color="0F172A")

instructions = [
    "1) Edita primero PLANES y REWARDS (catálogos maestros).",
    "2) Da de alta CLIENTES → ID Cliente, nombre, contacto.",
    "3) Crea cada SUSCRIPCIÓN: ID Cliente + ID Plan + Fecha inicio + Día de cobro. El Plan, Precio y Periodicidad se autocompletan.",
    "4) Cada vez que cobres, registra una fila en PAGOS: Fecha, ID Sub y Monto. Cliente y Plan se autocompletan.",
    "5) ESTADO DE CUENTA muestra automáticamente: cuotas esperadas, total facturado, total pagado, saldo, próximo cobro y mora.",
    "6) Revisa ALERTAS para ver de un vistazo quién debe pagar pronto y quién está atrasado.",
    "7) Registra rewards canjeados en CANJES → calcula automáticamente el saldo de puntos disponibles.",
    "8) DASHBOARD recalcula MRR, ARR, mora y puntos al abrir el archivo.",
    "",
    "🎨 Convención de color:  AZUL = entrada manual  ·  NEGRO = fórmula automática (no editar)",
    "🔄 Las celdas con TODAY() se recalculan al abrir Excel/Google Sheets.",
]
for i, t_ in enumerate(instructions):
    rr = 33 + i
    ws_db.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
    c = ws_db.cell(row=rr, column=1, value=t_)
    c.font = font(size=10, color="374151")
    c.alignment = Alignment(horizontal="left", wrap_text=True, indent=1)
    ws_db.row_dimensions[rr].height = 20

# Reordenar hojas
order = ["Dashboard", "Alertas", "EstadoCuenta", "Suscripciones", "Pagos", "Clientes", "Planes", "Rewards", "Canjes"]
wb._sheets = [wb[name] for name in order if name in wb.sheetnames]

# Freeze panes en hojas con datos
for ws_name in ["Suscripciones","Pagos","Clientes","Planes","Rewards","Canjes","EstadoCuenta"]:
    wb[ws_name].freeze_panes = "A5"
wb["Dashboard"].freeze_panes = "A4"
wb["Alertas"].freeze_panes  = "A4"

wb.save(OUT)
print(f"✓ Archivo creado: {OUT}")
print(f"✓ Hojas: {[ws.title for ws in wb.worksheets]}")
