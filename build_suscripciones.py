"""
Genera Suscripciones_Rewards.xlsx con 5 cuadros en una sola hoja.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "/Users/Bryan/mate_ai/mateAIweb/Suscripciones_Rewards.xlsx"

# --- Estilos ---
FONT = "Arial"
TITLE_FILL = PatternFill("solid", start_color="1F2937")
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="FFFFFF")
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="6B7280")

CUADRO_FILL = PatternFill("solid", start_color="2563EB")
CUADRO_FONT = Font(name=FONT, bold=True, size=12, color="FFFFFF")

HEADER_FILL = PatternFill("solid", start_color="111827")
HEADER_FONT = Font(name=FONT, bold=True, size=10, color="FFFFFF")

INPUT_FONT = Font(name=FONT, size=10, color="1D4ED8")        # azul = entradas
FORMULA_FONT = Font(name=FONT, size=10, color="000000")       # negro = fórmulas
NOTE_FONT = Font(name=FONT, italic=True, size=9, color="6B7280")

THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

KPI_LABEL_FILL = PatternFill("solid", start_color="F3F4F6")
KPI_VALUE_FILL = PatternFill("solid", start_color="FEF3C7")

wb = Workbook()
ws = wb.active
ws.title = "Suscripciones"
ws.sheet_view.showGridLines = False

# Anchos de columna
widths = {
    "A": 18, "B": 22, "C": 20, "D": 22, "E": 18, "F": 16, "G": 12,
    "H": 14, "I": 14, "J": 18, "K": 14, "L": 14, "M": 16, "N": 12,
    "O": 18, "P": 30
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

def cuadro_header(row, text, span=16):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = CUADRO_FILL
    c.font = CUADRO_FONT
    c.alignment = LEFT
    ws.row_dimensions[row].height = 26

def headers(row, cols):
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 32

def write_row(row, values, start_col=1, font=INPUT_FONT, align=LEFT, fmt=None):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.font = font
        c.alignment = align
        c.border = BORDER
        if fmt:
            c.number_format = fmt

# =========================================================
# TÍTULO
# =========================================================
ws.merge_cells("A1:P1")
t = ws.cell(row=1, column=1, value="SUSCRIPCIONES & REWARDS")
t.fill = TITLE_FILL
t.font = TITLE_FONT
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 38

ws.merge_cells("A2:P2")
s = ws.cell(row=2, column=1, value="Plataforma de fidelización por suscripción · MateAI · 5 cuadros vinculados")
s.font = SUBTITLE_FONT
s.alignment = Alignment(horizontal="left", indent=1)

# =========================================================
# CUADRO 1 — CATÁLOGO DE PLANES (row 4)
# Columnas: A ID Plan | B Nombre | C Periodicidad | D Precio | E Puntos/ciclo | F Rewards incluidos | G Activo
# =========================================================
cuadro_header(4, "▶ CUADRO 1 — CATÁLOGO DE PLANES")
plan_headers = ["ID Plan", "Nombre", "Periodicidad", "Precio", "Puntos por ciclo",
                "Rewards incluidos", "Activo"]
headers(5, plan_headers)

planes = [
    ("PLAN-BASIC",  "Basic",     "Mensual", 40,  50,   "Soporte estándar · 1 reward/mes",                 "Sí"),
    ("PLAN-PRO",    "Pro",       "Mensual", 80,  100,  "Soporte prioritario · 2 rewards · 5% desc. upsell","Sí"),
    ("PLAN-ELITE",  "Elite",     "Mensual", 150, 200,  "Soporte 24/7 · rewards ilimitados · 10% desc.",    "Sí"),
    ("PLAN-ANUAL",  "Pro Anual", "Anual",   800, 1500, "Plan Pro pagado anual · 2 meses gratis",          "Sí"),
]
for i, p in enumerate(planes):
    r = 6 + i
    write_row(r, p[:3])
    ws.cell(row=r, column=4, value=p[3]).number_format = '"$"#,##0'
    ws.cell(row=r, column=5, value=p[4]).number_format = '#,##0'
    ws.cell(row=r, column=6, value=p[5])
    ws.cell(row=r, column=7, value=p[6])
    for col in range(1, 8):
        cc = ws.cell(row=r, column=col)
        cc.font = INPUT_FONT
        cc.border = BORDER
        cc.alignment = LEFT if col != 4 and col != 5 else RIGHT

# Filas vacías reservadas (10–13) para futuros planes
for r in range(10, 14):
    for col in range(1, 8):
        ws.cell(row=r, column=col).border = BORDER

# Rango con nombre para VLOOKUP de planes
from openpyxl.workbook.defined_name import DefinedName
wb.defined_names["PLANES"] = DefinedName("PLANES", attr_text="Suscripciones!$A$6:$G$13")

# =========================================================
# CUADRO 2 — SUSCRIPCIONES ACTIVAS (row 15)
# Cols: A ID Sub | B ID Cliente | C Cliente | D ID Plan | E Plan(auto) | F Period(auto)
#       G Precio(auto) | H F.Inicio | I Ciclos pagados | J F.Próx renov(auto) | K Días rest(auto)
#       L Estado(auto) | M Método pago | N Auto-renov | O Puntos acum(auto) | P Notas
# =========================================================
cuadro_header(15, "▶ CUADRO 2 — SUSCRIPCIONES ACTIVAS")
sub_headers = ["ID Sub", "ID Cliente", "Cliente", "ID Plan", "Plan (auto)",
               "Periodicidad (auto)", "Precio (auto)", "Fecha inicio",
               "Ciclos pagados", "F. próx. renovación", "Días restantes",
               "Estado", "Método pago", "Auto-renov",
               "Puntos acumulados", "Notas"]
headers(16, sub_headers)

import datetime
hoy = datetime.date(2026, 5, 2)
suscripciones = [
    # ID Sub, ID Cli, Cliente, ID Plan, Fecha inicio, Ciclos, Método, Auto-renov, Notas
    ("SUB-0001", "CLI-001", "Jhon Amaya",   "PLAN-PRO",   datetime.date(2026, 3, 1), 2, "Transferencia", "Sí", "Cliente activo desde marzo"),
    ("SUB-0002", "CLI-002", "Restfull",     "PLAN-BASIC", datetime.date(2026, 4, 5), 1, "Transferencia", "Sí", ""),
    ("SUB-0003", "CLI-003", "Juli Riofrío", "PLAN-ELITE", datetime.date(2026, 1, 15), 4, "Tarjeta",        "Sí", "Cliente premium"),
    ("SUB-0004", "CLI-004", "Jean Aguilar", "PLAN-PRO",   datetime.date(2026, 4, 20), 1, "Efectivo",       "No", "Pago manual mensual"),
    ("SUB-0005", "CLI-005", "Avenaza SA",   "PLAN-ANUAL", datetime.date(2026, 2, 1), 1, "Transferencia", "Sí", "Renueva en 2027-02-01"),
]

start_row = 17
for i, s in enumerate(suscripciones):
    r = start_row + i
    # Inputs (azul)
    ws.cell(row=r, column=1, value=s[0]).font = INPUT_FONT
    ws.cell(row=r, column=2, value=s[1]).font = INPUT_FONT
    ws.cell(row=r, column=3, value=s[2]).font = INPUT_FONT
    ws.cell(row=r, column=4, value=s[3]).font = INPUT_FONT
    ws.cell(row=r, column=8, value=s[4]).font = INPUT_FONT
    ws.cell(row=r, column=8).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=9, value=s[5]).font = INPUT_FONT
    ws.cell(row=r, column=13, value=s[6]).font = INPUT_FONT
    ws.cell(row=r, column=14, value=s[7]).font = INPUT_FONT
    ws.cell(row=r, column=16, value=s[8]).font = INPUT_FONT

    # Fórmulas (negro)
    # E: Plan name -> VLOOKUP(D, PLANES, 2, FALSE)
    ws.cell(row=r, column=5, value=f'=IFERROR(VLOOKUP(D{r},PLANES,2,FALSE),"")')
    # F: Periodicidad -> VLOOKUP col 3
    ws.cell(row=r, column=6, value=f'=IFERROR(VLOOKUP(D{r},PLANES,3,FALSE),"")')
    # G: Precio -> VLOOKUP col 4
    ws.cell(row=r, column=7, value=f'=IFERROR(VLOOKUP(D{r},PLANES,4,FALSE),"")')
    ws.cell(row=r, column=7).number_format = '"$"#,##0'
    # J: Fecha próx. renovación = H + (I * (30 si Mensual / 365 si Anual))
    ws.cell(row=r, column=10, value=f'=IF(H{r}="","",H{r}+I{r}*IF(F{r}="Anual",365,30))')
    ws.cell(row=r, column=10).number_format = "yyyy-mm-dd"
    # K: Días restantes
    ws.cell(row=r, column=11, value=f'=IF(J{r}="","",J{r}-TODAY())')
    ws.cell(row=r, column=11).number_format = '0;[Red]-0;-'
    # L: Estado
    ws.cell(row=r, column=12, value=(
        f'=IF(K{r}="","",'
        f'IF(K{r}<0,"Vencida",'
        f'IF(K{r}<=7,"Por vencer",'
        f'IF(N{r}="No","Sin renov. auto","Activa"))))'
    ))
    # O: Puntos acumulados = ciclos * puntos por ciclo (VLOOKUP col 5)
    ws.cell(row=r, column=15, value=f'=IFERROR(I{r}*VLOOKUP(D{r},PLANES,5,FALSE),0)')
    ws.cell(row=r, column=15).number_format = '#,##0'

    # Estilo + borde a toda la fila
    for col in range(1, 17):
        c = ws.cell(row=r, column=col)
        c.border = BORDER
        if c.font.color is None or c.font.color.rgb not in ("001D4ED8", "FF1D4ED8"):
            if col in (5, 6, 7, 10, 11, 12, 15):
                c.font = FORMULA_FONT
        c.alignment = LEFT if col not in (4,7,9,11,15) else (RIGHT if col != 4 else LEFT)

# Filas vacías reservadas (22-30) con bordes y fórmulas
for r in range(start_row + len(suscripciones), 31):
    for col in range(1, 17):
        ws.cell(row=r, column=col).border = BORDER
    # Pre-cargar fórmulas para que funcionen al pegar datos
    ws.cell(row=r, column=5, value=f'=IFERROR(VLOOKUP(D{r},PLANES,2,FALSE),"")').font = FORMULA_FONT
    ws.cell(row=r, column=6, value=f'=IFERROR(VLOOKUP(D{r},PLANES,3,FALSE),"")').font = FORMULA_FONT
    cg = ws.cell(row=r, column=7, value=f'=IFERROR(VLOOKUP(D{r},PLANES,4,FALSE),"")')
    cg.font = FORMULA_FONT
    cg.number_format = '"$"#,##0'
    cj = ws.cell(row=r, column=10, value=f'=IF(H{r}="","",H{r}+I{r}*IF(F{r}="Anual",365,30))')
    cj.font = FORMULA_FONT
    cj.number_format = "yyyy-mm-dd"
    ck = ws.cell(row=r, column=11, value=f'=IF(J{r}="","",J{r}-TODAY())')
    ck.font = FORMULA_FONT
    ck.number_format = '0;[Red]-0;-'
    ws.cell(row=r, column=12, value=(
        f'=IF(K{r}="","",'
        f'IF(K{r}<0,"Vencida",'
        f'IF(K{r}<=7,"Por vencer",'
        f'IF(N{r}="No","Sin renov. auto","Activa"))))'
    )).font = FORMULA_FONT
    co = ws.cell(row=r, column=15, value=f'=IFERROR(I{r}*VLOOKUP(D{r},PLANES,5,FALSE),0)')
    co.font = FORMULA_FONT
    co.number_format = '#,##0'

# Formato condicional para columna L (Estado)
red_fill = PatternFill("solid", start_color="FECACA")
yellow_fill = PatternFill("solid", start_color="FEF3C7")
green_fill = PatternFill("solid", start_color="DCFCE7")
gray_fill = PatternFill("solid", start_color="E5E7EB")

ws.conditional_formatting.add(f"L17:L30",
    FormulaRule(formula=[f'$L17="Vencida"'], fill=red_fill, font=Font(name=FONT, bold=True, color="991B1B")))
ws.conditional_formatting.add(f"L17:L30",
    FormulaRule(formula=[f'$L17="Por vencer"'], fill=yellow_fill, font=Font(name=FONT, bold=True, color="92400E")))
ws.conditional_formatting.add(f"L17:L30",
    FormulaRule(formula=[f'$L17="Activa"'], fill=green_fill, font=Font(name=FONT, bold=True, color="166534")))
ws.conditional_formatting.add(f"L17:L30",
    FormulaRule(formula=[f'$L17="Sin renov. auto"'], fill=gray_fill, font=Font(name=FONT, bold=True, color="374151")))

# Validaciones de datos en cuadro 2
dv_plan = DataValidation(type="list", formula1="=$A$6:$A$13", allow_blank=True)
dv_plan.add(f"D17:D30")
ws.add_data_validation(dv_plan)

dv_yesno = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
dv_yesno.add(f"N17:N30")
ws.add_data_validation(dv_yesno)

dv_metodo = DataValidation(type="list", formula1='"Transferencia,Tarjeta,Efectivo,PayPal,Otro"', allow_blank=True)
dv_metodo.add(f"M17:M30")
ws.add_data_validation(dv_metodo)

# =========================================================
# CUADRO 3 — CATÁLOGO DE REWARDS (row 33)
# Cols: A ID Reward | B Nombre | C Tipo | D Costo puntos | E Plan mínimo | F Descripción | G Activo
# =========================================================
cuadro_header(33, "▶ CUADRO 3 — CATÁLOGO DE REWARDS")
rw_headers = ["ID Reward", "Nombre", "Tipo", "Costo puntos", "Plan mínimo",
              "Descripción", "Activo"]
headers(34, rw_headers)

rewards = [
    ("RW-001", "Descuento 10%",        "Descuento",   100, "PLAN-BASIC", "10% off en próximo upsell",       "Sí"),
    ("RW-002", "Descuento 25%",        "Descuento",   250, "PLAN-PRO",   "25% off en próximo upsell",       "Sí"),
    ("RW-003", "Mes gratis",           "Suscripción", 500, "PLAN-PRO",   "1 mes gratis al renovar",         "Sí"),
    ("RW-004", "Consultoría 30 min",   "Servicio",    300, "PLAN-PRO",   "Sesión 1:1 con especialista",     "Sí"),
    ("RW-005", "Consultoría 60 min",   "Servicio",    600, "PLAN-ELITE", "Sesión estratégica 60 min",       "Sí"),
    ("RW-006", "Acceso early features","Acceso",      150, "PLAN-BASIC", "Beta de nuevas funciones",        "Sí"),
    ("RW-007", "Branding pack",        "Producto",    400, "PLAN-PRO",   "Plantillas de marca exclusivas",  "Sí"),
    ("RW-008", "Auditoría completa",   "Servicio",   1000, "PLAN-ELITE", "Auditoría de negocio integral",   "Sí"),
]
for i, r_ in enumerate(rewards):
    rr = 35 + i
    ws.cell(row=rr, column=1, value=r_[0]).font = INPUT_FONT
    ws.cell(row=rr, column=2, value=r_[1]).font = INPUT_FONT
    ws.cell(row=rr, column=3, value=r_[2]).font = INPUT_FONT
    cd = ws.cell(row=rr, column=4, value=r_[3]); cd.font = INPUT_FONT; cd.number_format = "#,##0"
    ws.cell(row=rr, column=5, value=r_[4]).font = INPUT_FONT
    ws.cell(row=rr, column=6, value=r_[5]).font = INPUT_FONT
    ws.cell(row=rr, column=7, value=r_[6]).font = INPUT_FONT
    for col in range(1, 8):
        c = ws.cell(row=rr, column=col)
        c.border = BORDER
        c.alignment = LEFT if col != 4 else RIGHT

# Filas reservadas
for rr in range(43, 46):
    for col in range(1, 8):
        ws.cell(row=rr, column=col).border = BORDER

wb.defined_names["REWARDS"] = DefinedName("REWARDS", attr_text="Suscripciones!$A$35:$G$45")

# =========================================================
# CUADRO 4 — REWARDS ASIGNADOS / CANJEADOS (row 47)
# Cols: A ID Asign | B Fecha | C ID Cliente | D Cliente(auto) | E ID Sub
#       F ID Reward | G Reward(auto) | H Costo puntos(auto) | I Origen | J Estado
# =========================================================
cuadro_header(47, "▶ CUADRO 4 — REWARDS ASIGNADOS / CANJEADOS")
asg_headers = ["ID Asign", "Fecha", "ID Cliente", "Cliente (auto)", "ID Sub",
               "ID Reward", "Reward (auto)", "Costo puntos (auto)",
               "Origen", "Estado"]
headers(48, asg_headers)

# Construimos también un rango de clientes desde Cuadro 2 para lookup (B=ID Cli, C=Cliente)
wb.defined_names["CLIENTES_SUB"] = DefinedName("CLIENTES_SUB", attr_text="Suscripciones!$B$17:$C$30")

asignados = [
    ("ASG-0001", datetime.date(2026,4,10), "CLI-001", "SUB-0001", "RW-001", "Suscripción", "Canjeado"),
    ("ASG-0002", datetime.date(2026,4,15), "CLI-003", "SUB-0003", "RW-005", "Suscripción", "Canjeado"),
    ("ASG-0003", datetime.date(2026,4,22), "CLI-003", "SUB-0003", "RW-007", "Manual",       "Canjeado"),
    ("ASG-0004", datetime.date(2026,4,28), "CLI-001", "SUB-0001", "RW-006", "Automático",  "Pendiente"),
    ("ASG-0005", datetime.date(2026,5,1),  "CLI-005", "SUB-0005", "RW-002", "Suscripción", "Canjeado"),
]
for i, a in enumerate(asignados):
    rr = 49 + i
    ws.cell(row=rr, column=1, value=a[0]).font = INPUT_FONT
    cb = ws.cell(row=rr, column=2, value=a[1]); cb.font = INPUT_FONT; cb.number_format = "yyyy-mm-dd"
    ws.cell(row=rr, column=3, value=a[2]).font = INPUT_FONT
    # Cliente (auto) — lookup contra CLIENTES_SUB
    ws.cell(row=rr, column=4, value=f'=IFERROR(VLOOKUP(C{rr},CLIENTES_SUB,2,FALSE),"")').font = FORMULA_FONT
    ws.cell(row=rr, column=5, value=a[3]).font = INPUT_FONT
    ws.cell(row=rr, column=6, value=a[4]).font = INPUT_FONT
    # Reward (auto)
    ws.cell(row=rr, column=7, value=f'=IFERROR(VLOOKUP(F{rr},REWARDS,2,FALSE),"")').font = FORMULA_FONT
    # Costo puntos (auto)
    ch = ws.cell(row=rr, column=8, value=f'=IFERROR(VLOOKUP(F{rr},REWARDS,4,FALSE),"")')
    ch.font = FORMULA_FONT; ch.number_format = "#,##0"
    ws.cell(row=rr, column=9, value=a[5]).font = INPUT_FONT
    ws.cell(row=rr, column=10, value=a[6]).font = INPUT_FONT
    for col in range(1, 11):
        c = ws.cell(row=rr, column=col)
        c.border = BORDER
        c.alignment = LEFT if col not in (2, 8) else RIGHT

# Filas reservadas con fórmulas
for rr in range(54, 60):
    for col in range(1, 11):
        ws.cell(row=rr, column=col).border = BORDER
    ws.cell(row=rr, column=4, value=f'=IFERROR(VLOOKUP(C{rr},CLIENTES_SUB,2,FALSE),"")').font = FORMULA_FONT
    ws.cell(row=rr, column=7, value=f'=IFERROR(VLOOKUP(F{rr},REWARDS,2,FALSE),"")').font = FORMULA_FONT
    ch = ws.cell(row=rr, column=8, value=f'=IFERROR(VLOOKUP(F{rr},REWARDS,4,FALSE),"")')
    ch.font = FORMULA_FONT; ch.number_format = "#,##0"

# Validaciones cuadro 4
dv_origen = DataValidation(type="list", formula1='"Suscripción,Manual,Automático,Promoción"', allow_blank=True)
dv_origen.add("I49:I60")
ws.add_data_validation(dv_origen)

dv_estado = DataValidation(type="list", formula1='"Pendiente,Canjeado,Expirado,Cancelado"', allow_blank=True)
dv_estado.add("J49:J60")
ws.add_data_validation(dv_estado)

dv_reward = DataValidation(type="list", formula1="=$A$35:$A$45", allow_blank=True)
dv_reward.add("F49:F60")
ws.add_data_validation(dv_reward)

# Formato condicional para columna J Estado
ws.conditional_formatting.add("J49:J60",
    FormulaRule(formula=['$J49="Canjeado"'], fill=green_fill, font=Font(name=FONT, bold=True, color="166534")))
ws.conditional_formatting.add("J49:J60",
    FormulaRule(formula=['$J49="Pendiente"'], fill=yellow_fill, font=Font(name=FONT, bold=True, color="92400E")))
ws.conditional_formatting.add("J49:J60",
    FormulaRule(formula=['$J49="Expirado"'], fill=red_fill, font=Font(name=FONT, bold=True, color="991B1B")))

# =========================================================
# CUADRO 5 — KPIs (row 62)
# =========================================================
cuadro_header(62, "▶ CUADRO 5 — KPIs", span=4)
# Headers KPI
for i, h in enumerate(["KPI", "Valor", "Detalle / Fórmula"]):
    c = ws.cell(row=63, column=1+i, value=h)
    c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
ws.row_dimensions[63].height = 28
ws.column_dimensions["A"].width = 32
ws.column_dimensions["C"].width = 50

kpis = [
    ("Suscripciones activas",
     '=COUNTIF(L17:L30,"Activa")+COUNTIF(L17:L30,"Por vencer")+COUNTIF(L17:L30,"Sin renov. auto")',
     "Conteo de filas con estado distinto de Vencida", "#,##0"),
    ("Suscripciones por vencer (≤7 días)",
     '=COUNTIF(L17:L30,"Por vencer")',
     "Estado = Por vencer", "#,##0"),
    ("Suscripciones vencidas",
     '=COUNTIF(L17:L30,"Vencida")',
     "Estado = Vencida (acción: contactar y reactivar)", "#,##0"),
    ("Sin renovación automática",
     '=COUNTIF(L17:L30,"Sin renov. auto")',
     "Activas pero requieren cobro manual", "#,##0"),
    ("MRR — Ingreso recurrente mensual",
     '=SUMPRODUCT((F17:F30="Mensual")*(G17:G30))+SUMPRODUCT((F17:F30="Anual")*(G17:G30)/12)',
     "Suma precios mensuales + (anuales/12)", '"$"#,##0'),
    ("Ingreso anualizado (ARR)",
     '=B66*12',
     "MRR × 12", '"$"#,##0'),
    ("Total puntos otorgados",
     '=SUM(O17:O30)',
     "Suma de puntos acumulados de todas las suscripciones", "#,##0"),
    ("Total rewards canjeados",
     '=COUNTIF(J49:J60,"Canjeado")',
     "Cuadro 4 — Estado = Canjeado", "#,##0"),
    ("Puntos gastados en rewards",
     '=SUMIF(J49:J60,"Canjeado",H49:H60)',
     "Suma de costo en puntos de rewards canjeados", "#,##0"),
    ("Plan más popular",
     '=IFERROR(INDEX(B17:B30,MATCH(MAX(COUNTIF(E17:E30,E17:E30)),COUNTIF(E17:E30,E17:E30),0)),"")',
     "Plan con más suscripciones activas", None),
]

for i, (label, formula, detail, fmt) in enumerate(kpis):
    r = 64 + i
    cl = ws.cell(row=r, column=1, value=label)
    cl.fill = KPI_LABEL_FILL
    cl.font = Font(name=FONT, bold=True, size=10, color="111827")
    cl.alignment = LEFT
    cl.border = BORDER

    cv = ws.cell(row=r, column=2, value=formula)
    cv.fill = KPI_VALUE_FILL
    cv.font = Font(name=FONT, bold=True, size=11, color="111827")
    cv.alignment = RIGHT
    cv.border = BORDER
    if fmt:
        cv.number_format = fmt

    cd = ws.cell(row=r, column=3, value=detail)
    cd.font = NOTE_FONT
    cd.alignment = LEFT
    cd.border = BORDER
    ws.row_dimensions[r].height = 22

# =========================================================
# Notas finales
# =========================================================
note_row = 64 + len(kpis) + 2
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=12)
n = ws.cell(row=note_row, column=1, value="📌 Cómo usar:")
n.font = Font(name=FONT, bold=True, size=11, color="111827")

instructions = [
    "1) Edita Cuadro 1 (Planes) y Cuadro 3 (Rewards) primero — son los catálogos.",
    "2) En Cuadro 2 escribe sólo: ID Sub, ID Cliente, Cliente, ID Plan, Fecha inicio, Ciclos pagados, Método pago, Auto-renov, Notas. El resto se calcula solo (Plan, Periodicidad, Precio, Próx. renovación, Días restantes, Estado, Puntos acumulados).",
    "3) En Cuadro 4 registra cada canje: Fecha, ID Cliente, ID Sub, ID Reward, Origen, Estado. El Cliente y Reward se completan por VLOOKUP.",
    "4) Los colores de Estado son automáticos: verde=Activa, amarillo=Por vencer, rojo=Vencida, gris=Sin renov. auto.",
    "5) Cuadro 5 (KPIs) recalcula MRR, ARR, vencimientos y puntos automáticamente.",
    "6) Las celdas en azul son entradas. Las negras son fórmulas — no las edites.",
]
for i, t_ in enumerate(instructions):
    rr = note_row + 1 + i
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=12)
    c = ws.cell(row=rr, column=1, value=t_)
    c.font = Font(name=FONT, size=10, color="374151")
    c.alignment = Alignment(horizontal="left", wrap_text=True, indent=1)
    ws.row_dimensions[rr].height = 22

# Freeze panes para que el título quede visible
ws.freeze_panes = "A4"

wb.save(OUT)
print(f"✓ Archivo creado: {OUT}")
